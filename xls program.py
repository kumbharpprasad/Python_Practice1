import openpyxl

data_file = openpyxl.load_workbook("data.xlsx")  # To load the xls file
product_list = data_file["Sheet1"]  # To open the sheet1

#Task 1 > total number of products per suplier & list the name of suppiler with that products.

product_per_supplier = {}
total_value_per_supplier = {}
product_with_less_inventory = {}

for row in range(2, product_list.max_row + 1):  #.max_row : last row -1, +1 : last row inclusive, beacuse range start from 0-n
    supplier_name = product_list.cell(row, 4).value #row:row number, 4:coloum number, .value = grab the cell contains
    inventory = product_list.cell(row, 2).value #row:row number, 2:coloum number, .value = grab the cell contains
    price = product_list.cell(row, 3).value #row:row number, 3:coloum number, .value = grab the cell contains
    product_number = product_list.cell(row, 1).value #to access the 1st row.
    inventory_price = product_list.cell(row, 5) # to write the content on 5th coloum we need whole cell object and not just value.

#calculation for product per supplier
    if supplier_name in product_per_supplier:
        product_per_supplier[supplier_name] = product_per_supplier[supplier_name] + 1
    else:
        product_per_supplier[supplier_name] = 1
        print(f"New suplier {supplier_name} added to product record")

#Task2> calculation of total inventory value per supplier

    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)   #.get will show the value of key, here it is price.
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price
        print(f"New Supplier {supplier_name} added to value record")

#task3 > List all product with inventory under 10
    if inventory < 10:
        product_with_less_inventory[product_number] = inventory

#task4 > add new field to excel file; add value to total inventory price
    inventory_price.value = inventory * price    #.value here to set the value instead just grabing it

data_file.save("new_date_file.xlsx")   #To save as new excel file(not sheet).

print(product_per_supplier)
print(total_value_per_supplier)
print(product_with_less_inventory)
print ("New file has been created with name new_data_file")